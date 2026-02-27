"""
export_results.py — экспорт итогов конкурса в Excel

Структура файла:
- Лист "Общий отчёт"  — сводка по всем дням + статистика жюри
- Лист "День 1", "День 2", ... — принятые посты по каждому дню

Только принятые посты: HumanWords IS NOT NULL AND Rejected = 0

Запуск: python scripts/export_results.py
"""

import sys
import pathlib
import sqlite3
from datetime import datetime

ROOT = pathlib.Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

DB_PATH      = ROOT / "data" / "main.db"
RESULTS_DIR  = ROOT / "results"

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Не установлен openpyxl. Выполни: pip install openpyxl")
    sys.exit(1)


# ── Цвета ─────────────────────────────────────────────────────────────────────

CLR_HEADER      = "2C3E50"   # тёмно-синий — заголовки таблиц
CLR_SUBHEADER   = "5D6D7E"   # серый — подзаголовки
CLR_ACCENT      = "3498DB"   # синий — выделение
CLR_LIGHT_ROW   = "EBF5FB"   # светло-голубой — чётные строки
CLR_WHITE       = "FFFFFF"
CLR_FONT_LIGHT  = "FFFFFF"
CLR_FONT_DARK   = "1A1A1A"
CLR_GREEN       = "1E8449"
CLR_SECTION     = "D6EAF8"   # светлый фон секций


def _header_font(size=11, bold=True, color=CLR_FONT_LIGHT):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _body_font(size=10, bold=False, color=CLR_FONT_DARK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="BDC3C7")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row, cols, bg=CLR_HEADER):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _header_font()
        cell.fill      = _fill(bg)
        cell.alignment = _center()
        cell.border    = _border()


def _style_data_row(ws, row, cols, even=False):
    bg = CLR_LIGHT_ROW if even else CLR_WHITE
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _body_font()
        cell.fill      = _fill(bg)
        cell.alignment = _left() if col <= 3 else _center()
        cell.border    = _border()


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Данные из БД ──────────────────────────────────────────────────────────────

def _connect() -> sqlite3.Connection:
    if not DB_PATH.exists():
        print(f"❌ База данных не найдена: {DB_PATH}")
        print("   Запусти сначала: python scripts/init_db.py")
        sys.exit(1)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _get_days(conn) -> list:
    return conn.execute("SELECT Day, Data FROM days ORDER BY Day").fetchall()


def _get_posts_by_day(conn, day_id: int) -> list:
    """Принятые посты за день: HumanWords NOT NULL, Rejected = 0."""
    return conn.execute(
        """
        SELECT
            a.Name          AS author,
            a.URL           AS author_url,
            p.URL           AS post_url,
            r.BotWords      AS bot_words,
            r.HumanWords    AS human_words,
            r.HumanErrors   AS human_errors,
            rv.Name         AS reviewer
        FROM posts_info p
        JOIN authors  a  ON p.Author   = a.ID
        JOIN results  r  ON r.Post     = p.ID
        LEFT JOIN reviewers rv ON r.Reviewer = rv.TGID
        WHERE p.Day          = ?
          AND p.Rejected      = 0
          AND p.PostOfReviewer = 0
          AND r.HumanWords   IS NOT NULL
        ORDER BY a.Name
        """,
        (day_id,),
    ).fetchall()


def _get_summary_by_day(conn) -> list:
    return conn.execute(
        """
        SELECT
            d.Day,
            d.Data                                          AS date,
            COUNT(p.ID)                                     AS total,
            COALESCE(SUM(r.HumanWords), 0)                 AS total_words,
            COALESCE(SUM(r.HumanErrors), 0)                AS total_errors
        FROM days d
        LEFT JOIN posts_info p ON p.Day = d.Day
            AND p.Rejected = 0
            AND p.PostOfReviewer = 0
        LEFT JOIN results r ON r.Post = p.ID
            AND r.HumanWords IS NOT NULL
        GROUP BY d.Day
        ORDER BY d.Day
        """,
    ).fetchall()


def _get_reviewer_stats(conn) -> list:
    return conn.execute(
        """
        SELECT
            rv.Name,
            COUNT(r.ID)                         AS checked,
            COALESCE(SUM(r.HumanWords), 0)      AS total_words,
            COALESCE(SUM(r.HumanErrors), 0)     AS total_errors
        FROM reviewers rv
        LEFT JOIN results r ON r.Reviewer = rv.TGID
            AND r.HumanWords IS NOT NULL
        WHERE rv.Verified = 1
        GROUP BY rv.TGID
        ORDER BY checked DESC
        """,
    ).fetchall()


def _get_totals(conn) -> dict:
    row = conn.execute(
        """
        SELECT
            COUNT(p.ID)                         AS posts,
            COALESCE(SUM(r.HumanWords), 0)      AS words,
            COALESCE(SUM(r.HumanErrors), 0)     AS errors
        FROM posts_info p
        JOIN results r ON r.Post = p.ID
        WHERE p.Rejected = 0
          AND p.PostOfReviewer = 0
          AND r.HumanWords IS NOT NULL
        """
    ).fetchone()
    return dict(row)


# ── Лист: Общий отчёт ─────────────────────────────────────────────────────────

def _build_summary_sheet(ws, conn):
    ws.title = "Общий отчёт"
    ws.sheet_view.showGridLines = False

    # ── Заголовок ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value     = "ИТОГИ КОНКУРСА inkstory.net"
    title.font      = Font(name="Arial", size=16, bold=True, color=CLR_FONT_LIGHT)
    title.fill      = _fill(CLR_HEADER)
    title.alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value     = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sub.font      = Font(name="Arial", size=10, color=CLR_FONT_LIGHT)
    sub.fill      = _fill(CLR_SUBHEADER)
    sub.alignment = _center()
    ws.row_dimensions[2].height = 20

    # ── Итого по конкурсу ─────────────────────────────────────────────────────
    totals = _get_totals(conn)
    errors_per_1000 = round(totals["errors"] / totals["words"] * 1000, 2) if totals["words"] else 0

    ws.row_dimensions[3].height = 10  # отступ

    ws.merge_cells("A4:G4")
    sec = ws["A4"]
    sec.value     = "ОБЩИЕ ПОКАЗАТЕЛИ"
    sec.font      = _header_font(size=11)
    sec.fill      = _fill(CLR_ACCENT)
    sec.alignment = _center()
    ws.row_dimensions[4].height = 24

    summary_data = [
        ("Принято постов",          totals["posts"]),
        ("Всего слов (по жюри)",     totals["words"]),
        ("Всего ошибок",             totals["errors"]),
        ("Ошибок на 1000 слов",      errors_per_1000),
    ]

    for i, (label, value) in enumerate(summary_data, start=5):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=1).font  = _body_font(bold=True)
        ws.cell(row=i, column=1).fill  = _fill(CLR_SECTION)
        ws.cell(row=i, column=1).alignment = _left()
        ws.cell(row=i, column=1).border = _border()

        ws.cell(row=i, column=2).value = value
        ws.cell(row=i, column=2).font  = _body_font(bold=True, color=CLR_GREEN)
        ws.cell(row=i, column=2).fill  = _fill(CLR_WHITE)
        ws.cell(row=i, column=2).alignment = _center()
        ws.cell(row=i, column=2).border = _border()

        for col in range(3, 8):
            ws.cell(row=i, column=col).fill   = _fill(CLR_WHITE)
            ws.cell(row=i, column=col).border = _border()

        ws.row_dimensions[i].height = 22

    # ── По дням ───────────────────────────────────────────────────────────────
    r = 10
    ws.row_dimensions[r - 1].height = 10

    ws.merge_cells(f"A{r}:G{r}")
    sec2 = ws[f"A{r}"]
    sec2.value     = "ПО ДНЯМ КОНКУРСА"
    sec2.font      = _header_font(size=11)
    sec2.fill      = _fill(CLR_ACCENT)
    sec2.alignment = _center()
    ws.row_dimensions[r].height = 24

    r += 1
    headers = ["День", "Дата", "Принято постов", "Слов (жюри)", "Ошибок", "Ошибок / 1000 слов", ""]
    for col, h in enumerate(headers, 1):
        ws.cell(row=r, column=col).value = h
    _style_header_row(ws, r, 6)
    ws.row_dimensions[r].height = 22

    days_data = _get_summary_by_day(conn)
    for i, day in enumerate(days_data, start=1):
        r += 1
        ep1k = round(day["total_errors"] / day["total_words"] * 1000, 2) if day["total_words"] else 0
        row_data = [
            day["Day"], day["date"], day["total"],
            day["total_words"], day["total_errors"], ep1k,
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=r, column=col).value = val
        _style_data_row(ws, r, 6, even=(i % 2 == 0))
        ws.row_dimensions[r].height = 20

    # Итого строка
    r += 1
    first_data = 12
    ws.cell(row=r, column=1).value = "ИТОГО"
    ws.cell(row=r, column=1).font  = _header_font()
    for col in range(1, 7):
        c = ws.cell(row=r, column=col)
        c.fill   = _fill(CLR_HEADER)
        c.font   = _header_font()
        c.border = _border()
        c.alignment = _center()
    ws.cell(row=r, column=3).value = f"=SUM(C{first_data}:C{r-1})"
    ws.cell(row=r, column=4).value = f"=SUM(D{first_data}:D{r-1})"
    ws.cell(row=r, column=5).value = f"=SUM(E{first_data}:E{r-1})"
    ws.cell(row=r, column=6).value = f"=IFERROR(ROUND(E{r}/D{r}*1000,2),0)"
    ws.row_dimensions[r].height = 22

    # ── Статистика жюри ───────────────────────────────────────────────────────
    r += 2
    ws.merge_cells(f"A{r}:G{r}")
    sec3 = ws[f"A{r}"]
    sec3.value     = "СТАТИСТИКА ЖЮРИ"
    sec3.font      = _header_font(size=11)
    sec3.fill      = _fill(CLR_ACCENT)
    sec3.alignment = _center()
    ws.row_dimensions[r].height = 24

    r += 1
    rev_headers = ["Жюри", "Проверено постов", "Слов проверено", "Ошибок найдено", "Ошибок / 1000 слов", "", ""]
    for col, h in enumerate(rev_headers, 1):
        ws.cell(row=r, column=col).value = h
    _style_header_row(ws, r, 5)
    ws.row_dimensions[r].height = 22

    rev_data = _get_reviewer_stats(conn)
    for i, rv in enumerate(rev_data, start=1):
        r += 1
        ep1k = round(rv["total_errors"] / rv["total_words"] * 1000, 2) if rv["total_words"] else 0
        row_data = [rv["Name"], rv["checked"], rv["total_words"], rv["total_errors"], ep1k]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=r, column=col).value = val
        _style_data_row(ws, r, 5, even=(i % 2 == 0))
        ws.row_dimensions[r].height = 20

    # ── Ширины колонок ────────────────────────────────────────────────────────
    _set_col_widths(ws, {
        "A": 20, "B": 16, "C": 18,
        "D": 18, "E": 16, "F": 22, "G": 5,
    })


# ── Лист: День ────────────────────────────────────────────────────────────────

def _build_day_sheet(ws, day_row, posts: list):
    ws.title = f"День {day_row['Day']}"
    ws.sheet_view.showGridLines = False

    # Заголовок
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value     = f"День {day_row['Day']}  —  {day_row['Data']}"
    title.font      = Font(name="Arial", size=14, bold=True, color=CLR_FONT_LIGHT)
    title.fill      = _fill(CLR_HEADER)
    title.alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value     = f"Принятых постов: {len(posts)}"
    sub.font      = Font(name="Arial", size=10, color=CLR_FONT_LIGHT)
    sub.fill      = _fill(CLR_SUBHEADER)
    sub.alignment = _center()
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8

    # Заголовки таблицы
    headers = ["#", "Автор", "Ссылка на пост", "Бот (слов)", "Жюри (слов)", "Ошибок", "Проверяющий"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = h
    _style_header_row(ws, 4, 7)
    ws.row_dimensions[4].height = 22

    if not posts:
        ws.merge_cells("A5:G5")
        empty = ws["A5"]
        empty.value     = "Нет данных за этот день"
        empty.font      = _body_font(color="888888")
        empty.alignment = _center()
        return

    for i, post in enumerate(posts, start=1):
        r = 4 + i
        row_data = [
            i,
            post["author"],
            post["post_url"],
            post["bot_words"],
            post["human_words"],
            post["human_errors"],
            post["reviewer"] or "—",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=r, column=col).value = val
        _style_data_row(ws, r, 7, even=(i % 2 == 0))
        ws.row_dimensions[r].height = 20

    # Итого
    last = 4 + len(posts)
    r = last + 1
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.fill   = _fill(CLR_HEADER)
        c.font   = _header_font()
        c.border = _border()
        c.alignment = _center()
    ws.cell(row=r, column=1).value = "ИТОГО"
    ws.cell(row=r, column=4).value = f"=SUM(D5:D{last})"
    ws.cell(row=r, column=5).value = f"=SUM(E5:E{last})"
    ws.cell(row=r, column=6).value = f"=SUM(F5:F{last})"
    ws.row_dimensions[r].height = 22

    # Ошибок на 1000 слов
    r2 = r + 1
    ws.merge_cells(f"A{r2}:C{r2}")
    label = ws[f"A{r2}"]
    label.value     = "Ошибок на 1000 слов:"
    label.font      = _body_font(bold=True)
    label.fill      = _fill(CLR_SECTION)
    label.alignment = _left()
    label.border    = _border()

    val_cell = ws[f"D{r2}"]
    val_cell.value     = f"=IFERROR(ROUND(F{r}/E{r}*1000,2),0)"
    val_cell.font      = _body_font(bold=True, color=CLR_GREEN)
    val_cell.fill      = _fill(CLR_SECTION)
    val_cell.alignment = _center()
    val_cell.border    = _border()

    for col in range(5, 8):
        c = ws.cell(row=r2, column=col)
        c.fill   = _fill(CLR_SECTION)
        c.border = _border()

    ws.row_dimensions[r2].height = 22

    _set_col_widths(ws, {
        "A": 5, "B": 22, "C": 40,
        "D": 14, "E": 14, "F": 12, "G": 20,
    })


# ── Точка входа ───────────────────────────────────────────────────────────────

def export() -> pathlib.Path:
    conn = _connect()

    days     = _get_days(conn)
    wb       = Workbook()

    # Удаляем дефолтный лист
    wb.remove(wb.active)

    # Лист 1 — общий отчёт
    ws_summary = wb.create_sheet("Общий отчёт")
    _build_summary_sheet(ws_summary, conn)

    # Листы по дням
    for day in days:
        posts = _get_posts_by_day(conn, day["Day"])
        ws_day = wb.create_sheet()
        _build_day_sheet(ws_day, day, posts)

    conn.close()

    # Сохраняем
    RESULTS_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_path  = RESULTS_DIR / f"results_{timestamp}.xlsx"
    wb.save(out_path)

    return out_path


def main():
    print("📊 Экспорт итогов конкурса...")
    out = export()
    print(f"✅ Готово: {out}")


if __name__ == "__main__":
    main()