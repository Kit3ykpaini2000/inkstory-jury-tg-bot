"""
cli.py — консольное управление inkstory-bot v3

Запуск: python cli.py
"""

import os
import sys
import pathlib
import sqlite3
from datetime import datetime

ROOT = pathlib.Path(__file__).parent
sys.path.insert(0, str(ROOT))

from utils.database import get_db
from utils.db.jury  import (
    is_admin, get_all_reviewers, set_verified, set_admin,
    delete_reviewer, get_reviewer_stats,
)
from utils.db.posts import get_posts_stats, errors_per_1000, release_stuck_posts
from utils.db.days  import get_all_days, create_day, delete_day

try:
    from parser.queue_manager import (
        get_total_queue_count, get_queue_count, assign_post,
        remove_post, release_post, get_all_reviewer_queue_sizes,
    )
    _QM = True
except ImportError:
    _QM = False

# ── Цвета ─────────────────────────────────────────────────────────────────────
R="\033[91m"; G="\033[92m"; Y="\033[93m"; B="\033[94m"
C="\033[96m"; W="\033[97m"; DIM="\033[2m"; BOLD="\033[1m"; RESET="\033[0m"

def clr(): os.system("cls" if os.name == "nt" else "clear")
def hr(char="─", color=DIM):
    try: w = os.get_terminal_size().columns
    except: w = 80
    print(f"{color}{char * w}{RESET}")
def pause(): input(f"\n{DIM}[ Enter — назад ]{RESET}")
def confirm(msg): return input(f"{Y}⚠  {msg} [y/N]: {RESET}").strip().lower() == "y"

def header(title, show_stats=False):
    clr(); hr("═", B)
    now = datetime.now().strftime("%d.%m.%Y  %H:%M")
    print(f"{BOLD}{C}  InkStory CLI v3  {DIM}│{RESET}{BOLD}  {title}  {DIM}│  {now}{RESET}")
    hr("═", B)
    if show_stats: print(_quick_stats()); hr()
    print()

def menu(title, options, show_stats=False):
    header(title, show_stats=show_stats)
    for i, opt in enumerate(options, 1):
        print(f"  {BOLD}{C}{i}{RESET}. {opt}")
    print(f"  {DIM}0. ← Назад / Выход{RESET}\n")
    while True:
        try:
            raw = input(f"{W}Выбор: {RESET}").strip()
            if raw == "0": return -1
            idx = int(raw) - 1
            if 0 <= idx < len(options): return idx
        except (ValueError, EOFError): pass
        print(f"{R}  Введите число от 0 до {len(options)}{RESET}")

def _quick_stats():
    try:
        s = get_posts_stats()
        in_queue = get_total_queue_count() if _QM else 0
        chk = f"  {Y}🔒 Проверяется: {s['checking']}{RESET}" if s['checking'] else ""
        return (f"  {B}📋 В очереди: {in_queue}{RESET}   "
                f"{G}✅ Проверено: {s['done']}{RESET}{chk}")
    except: return ""

def _status_label(status):
    return {
        "pending":       f"{Y}ожидает{RESET}",
        "checking":      f"{B}проверяется{RESET}",
        "done":          f"{G}проверен{RESET}",
        "rejected":      f"{R}отклонён{RESET}",
        "reviewer_post": f"{DIM}жюри-пост{RESET}",
    }.get(status, status)


# ══════════════════════════════════════════════════════════════════════════════
# СТАТИСТИКА
# ══════════════════════════════════════════════════════════════════════════════

def show_stats():
    header("📊 Статистика")
    s        = get_posts_stats()
    in_queue = get_total_queue_count() if _QM else 0
    rows     = get_reviewer_stats()

    print(f"  {Y}⏳ Ожидает проверки:{RESET}  {s['pending']}")
    print(f"  {Y}🔒 Проверяется:{RESET}        {s['checking']}")
    print(f"  {G}✅ Проверено:{RESET}          {s['done']}")
    print(f"  {R}❌ Отклонено:{RESET}          {s['rejected']}")
    print(f"  {W}📦 Всего постов:{RESET}       {s['total']}")
    print(f"  {B}📋 В очереди:{RESET}          {in_queue}")
    print(f"  {Y}🔗 Ссылок в БД:{RESET}        {s['links']}")

    if rows:
        print()
        hr()
        print(f"  {BOLD}Жюри — рейтинг:{RESET}")
        print(f"  {'Имя':<22} {'Проверено':>10} {'Отклонено':>10}")
        hr()
        for i, r in enumerate(rows, 1):
            bar = "█" * min(r["checked"], 30)
            print(f"  {i:2}. {r['Name']:<20} {C}{bar}{RESET} {r['checked']}  ❌{r['rejected']}")
    pause()


# ══════════════════════════════════════════════════════════════════════════════
# ЖЮРИ
# ══════════════════════════════════════════════════════════════════════════════

def manage_reviewers():
    while True:
        choice = menu("👥 Управление жюри", [
            "Список жюри", "Добавить жюри",
            "Верифицировать", "Снять верификацию",
            "Сделать администратором", "Снять права администратора",
            "Удалить жюри",
        ])
        if choice == -1: return
        [list_reviewers, add_reviewer,
         lambda: _set_ver(1), lambda: _set_ver(0),
         lambda: _set_adm(1), lambda: _set_adm(0),
         _delete_reviewer][choice]()

def list_reviewers():
    header("👥 Список жюри")
    rows = get_all_reviewers()
    if not rows:
        print(f"  {DIM}Жюри не зарегистрированы.{RESET}")
    else:
        print(f"  {'TGID':<15} {'Имя':<22} {'Верф':4} {'Адм':4} {'Провер':>7}")
        hr()
        for r in rows:
            adm = f"{Y}★{RESET}" if r["is_admin"] else " "
            ver = f"{G}✓{RESET}" if r["verified"] else f"{R}✗{RESET}"
            print(f"  {r['tgid']:<15} {r['name']:<22} {ver}    {adm}    {r['checked']:>5}")
    pause()

def _pick_reviewer(prompt="Выберите жюри"):
    rows = get_all_reviewers()
    if not rows:
        print(f"  {R}Жюри не найдены.{RESET}"); pause(); return None
    opts = [f"{r['name']}  (tgID: {r['tgid']})" for r in rows]
    idx  = menu(prompt, opts)
    if idx == -1: return None
    return rows[idx]["tgid"]

def add_reviewer():
    header("➕ Добавить жюри")
    name   = input(f"  Имя (никнейм): {W}").strip(); print(RESET, end="")
    tg_id  = input(f"  Telegram ID:   {W}").strip(); print(RESET, end="")
    url    = input(f"  URL профиля:   {W}").strip(); print(RESET, end="")
    is_adm = input(f"  Администратор? [y/N]: {W}").strip().lower() == "y"; print(RESET, end="")
    if not name or not tg_id:
        print(f"{R}  Имя и tgID обязательны.{RESET}"); pause(); return
    with get_db() as db:
        try:
            db.execute(
                "INSERT INTO reviewers (TGID, URL, Name, IsAdmin, Verified) VALUES (?,?,?,?,0)",
                (tg_id, url, name, 1 if is_adm else 0),
            )
            db.commit()
            print(f"\n{G}  ✅ Жюри {name} добавлен (не верифицирован).{RESET}")
        except sqlite3.IntegrityError:
            print(f"\n{R}  ❌ Такой tgID уже есть в БД.{RESET}")
    pause()

def _set_ver(value):
    tgid = _pick_reviewer("Верифицировать жюри" if value else "Снять верификацию")
    if not tgid: return
    set_verified(tgid, value)
    status = f"{G}верифицирован{RESET}" if value else f"{Y}верификация снята{RESET}"
    print(f"\n  ✅ Жюри tgID={tgid} {status}"); pause()

def _set_adm(value):
    tgid = _pick_reviewer("Назначить администратора" if value else "Снять права администратора")
    if not tgid: return
    set_admin(tgid, value)
    status = f"{G}назначен администратором{RESET}" if value else f"{Y}снят с администратора{RESET}"
    print(f"\n  ✅ Жюри tgID={tgid} {status}"); pause()

def _delete_reviewer():
    tgid = _pick_reviewer("❌ Удалить жюри")
    if not tgid: return
    rows = get_all_reviewers()
    name = next((r["name"] for r in rows if r["tgid"] == tgid), tgid)
    if not confirm(f"Удалить жюри «{name}»? Его результаты сохранятся."): return
    delete_reviewer(tgid)
    print(f"\n{G}  ✅ Жюри {name} удалён.{RESET}"); pause()


# ══════════════════════════════════════════════════════════════════════════════
# ПОСТЫ
# ══════════════════════════════════════════════════════════════════════════════

def manage_posts():
    while True:
        choice = menu("📝 Управление постами", [
            "Все посты", "Просмотр очереди",
            "Найти пост по URL", "Найти посты по автору",
            "Сбросить зависшие посты", "Сбросить проверку поста",
            "Отклонить пост вручную", "Восстановить отклонённый пост",
            "🔄 Переназначить очередь заново", "🗑 Очистить всю очередь",
        ])
        if choice == -1: return
        [view_all_posts, view_queue,
         find_post, find_post_by_author,
         release_stuck, reset_post,
         reject_post_cli, restore_post,
         reassign_queue, clear_queue][choice]()

def view_all_posts():
    header("📄 Все посты")
    status_opts = ["Все", "Ожидают", "Проверяются", "Проверены", "Отклонённые"]
    status_idx  = menu("Фильтр по статусу", status_opts)
    if status_idx == -1: return
    where = {
        0: "WHERE p.Status != 'reviewer_post'",
        1: "WHERE p.Status = 'pending'",
        2: "WHERE p.Status = 'checking'",
        3: "WHERE p.Status = 'done'",
        4: "WHERE p.Status = 'rejected'",
    }[status_idx]
    with get_db() as db:
        rows = db.execute(
            f"""
            SELECT p.ID, p.URL, p.Status, a.Name,
                   r.BotWords, r.HumanWords, r.HumanErrors, r.RejectReason
            FROM posts_info p
            JOIN authors    a ON p.Author = a.ID
            LEFT JOIN results r ON r.Post = p.ID
            {where}
            ORDER BY p.ID DESC LIMIT 50
            """
        ).fetchall()
    header(f"📄 Посты — {status_opts[status_idx]} (последние {len(rows)})")
    print(f"  {'#':<6} {'Автор':<22} {'Статус':<14} {'Бот сл':>6} {'Чел сл':>7} {'Ош':>4} {'Ош/1000':>8}")
    hr()
    for r in rows:
        ep   = errors_per_1000(r["HumanErrors"] or 0, r["HumanWords"] or 0) if r["HumanWords"] else "—"
        bw   = str(r["BotWords"]   or "—")
        hw   = str(r["HumanWords"] or "—")
        he   = str(r["HumanErrors"] or "—")
        print(f"  {r['ID']:<6} {r['Name']:<22} {_status_label(r['Status']):<22} {bw:>6} {hw:>7} {he:>4} {str(ep):>8}")
    pause()

def view_queue():
    header("📋 Очереди жюри")
    if not _QM:
        print(f"  {R}queue_manager недоступен.{RESET}"); pause(); return
    sizes = get_all_reviewer_queue_sizes()
    if not sizes:
        print(f"  {R}Нет верифицированных жюри.{RESET}"); pause(); return
    total = 0
    for rv in sizes:
        count = rv["count"]; total += count
        bar   = f"{B}{'█' * min(count, 20)}{RESET}" if count else f"{DIM}пусто{RESET}"
        print(f"  {rv['name']:<22} {bar} {count}")
    hr()
    print(f"  Всего в очередях: {total}")
    with get_db() as db:
        free = db.execute(
            "SELECT COUNT(*) FROM queue q JOIN posts_info p ON q.Post=p.ID "
            "WHERE q.Reviewer IS NULL AND p.Status='pending'"
        ).fetchone()[0]
    if free: print(f"  {DIM}Свободных (без жюри): {free}{RESET}")
    pause()

def find_post():
    header("🔍 Найти пост по URL")
    url = input(f"  URL или часть: {W}").strip(); print(RESET, end="")
    with get_db() as db:
        rows = db.execute(
            """
            SELECT p.ID, p.URL, p.Status, a.Name, r.BotWords
            FROM posts_info p
            JOIN authors    a ON p.Author = a.ID
            LEFT JOIN results r ON r.Post = p.ID
            WHERE p.URL LIKE ? LIMIT 10
            """, (f"%{url}%",)
        ).fetchall()
    if not rows:
        print(f"  {R}Не найдено.{RESET}")
    else:
        for r in rows:
            print(f"\n  {BOLD}#{r['ID']}{RESET} {r['Name']}")
            print(f"  {DIM}{r['URL']}{RESET}")
            print(f"  Статус: {_status_label(r['Status'])}  |  Бот: {r['BotWords'] or '—'} слов")
    pause()

def find_post_by_author():
    header("🔍 Найти посты по автору")
    name = input(f"  Имя автора или часть: {W}").strip(); print(RESET, end="")
    if not name: return
    with get_db() as db:
        rows = db.execute(
            """
            SELECT p.ID, p.URL, p.Status, a.Name,
                   r.BotWords, r.HumanWords, r.HumanErrors
            FROM posts_info p
            JOIN authors    a ON p.Author = a.ID
            LEFT JOIN results r ON r.Post = p.ID
            WHERE a.Name LIKE ?
            ORDER BY p.ID DESC LIMIT 20
            """, (f"%{name}%",)
        ).fetchall()
    if not rows:
        print(f"  {R}Не найдено.{RESET}"); pause(); return
    header(f"🔍 Посты «{name}» ({len(rows)} шт.)")
    print(f"  {'#':<6} {'Автор':<22} {'Статус':<14} {'Бот сл':>6} {'Чел сл':>7} {'Ош':>4}")
    hr()
    for r in rows:
        print(f"  {r['ID']:<6} {r['Name']:<22} {_status_label(r['Status']):<22} "
              f"{str(r['BotWords'] or '—'):>6} {str(r['HumanWords'] or '—'):>7} {str(r['HumanErrors'] or '—'):>4}")
    pause()

def release_stuck():
    header("🔓 Сброс зависших постов")
    with get_db() as db:
        rows = db.execute(
            """
            SELECT q.Post, p.URL, a.Name AS author,
                   rv.Name AS reviewer, q.TakenAt
            FROM queue q
            JOIN posts_info p  ON q.Post     = p.ID
            JOIN authors    a  ON p.Author   = a.ID
            LEFT JOIN reviewers rv ON rv.TGID = q.Reviewer
            WHERE q.TakenAt IS NOT NULL AND p.Status = 'checking'
            """
        ).fetchall()
    if not rows:
        print(f"  {G}Зависших постов нет!{RESET}"); pause(); return
    print(f"  {Y}Найдено: {len(rows)}{RESET}\n")
    print(f"  {'#':<6} {'Автор поста':<22} {'Взял жюри':<20}  Взято в"); hr()
    for r in rows:
        print(f"  {r['Post']:<6} {r['author']:<22} {r['reviewer'] or '—':<20}  {r['TakenAt']}")
    print()
    if not confirm(f"Освободить все {len(rows)} зависших постов?"): return
    count = release_stuck_posts()
    print(f"\n{G}  ✅ Освобождено {count} постов.{RESET}"); pause()

def reset_post():
    header("🔄 Сбросить проверку поста")
    raw = input(f"  ID поста: {W}").strip(); print(RESET, end="")
    if not raw.isdigit(): print(f"{R}  Нужен числовой ID.{RESET}"); pause(); return
    pid = int(raw)
    with get_db() as db:
        row = db.execute("SELECT URL, Status FROM posts_info WHERE ID=?", (pid,)).fetchone()
        if not row: print(f"{R}  Пост #{pid} не найден.{RESET}"); pause(); return
        if not confirm(f"Сбросить проверку поста #{pid} (статус: {row['Status']})?"): return
        db.execute("UPDATE posts_info SET Status='pending' WHERE ID=?", (pid,))
        db.execute("UPDATE queue SET Reviewer=NULL, TakenAt=NULL WHERE Post=?", (pid,))
        db.execute("UPDATE results SET HumanWords=NULL, HumanErrors=NULL, Reviewer=NULL, RejectReason=NULL WHERE Post=?", (pid,))
        db.commit()
    if _QM:
        with get_db() as db:
            in_q = db.execute("SELECT 1 FROM queue WHERE Post=?", (pid,)).fetchone()
        if not in_q: assign_post(pid)
    print(f"\n{G}  ✅ Пост #{pid} сброшен и возвращён в очередь.{RESET}"); pause()

def reject_post_cli():
    header("❌ Отклонить пост")
    raw = input(f"  ID поста: {W}").strip(); print(RESET, end="")
    if not raw.isdigit(): print(f"{R}  Нужен числовой ID.{RESET}"); pause(); return
    reason = input(f"  Причина: {W}").strip(); print(RESET, end="")
    pid = int(raw)
    with get_db() as db:
        row = db.execute("SELECT URL FROM posts_info WHERE ID=?", (pid,)).fetchone()
        if not row: print(f"{R}  Пост #{pid} не найден.{RESET}"); pause(); return
        if not confirm(f"Отклонить пост #{pid}?"): return
        db.execute("UPDATE posts_info SET Status='rejected' WHERE ID=?", (pid,))
        exists = db.execute("SELECT ID FROM results WHERE Post=?", (pid,)).fetchone()
        if exists:
            db.execute("UPDATE results SET Reviewer=NULL, RejectReason=? WHERE Post=?", (reason, pid))
        else:
            db.execute("INSERT INTO results (Post, RejectReason) VALUES (?,?)", (pid, reason))
        db.commit()
    if _QM: remove_post(pid)
    print(f"\n{G}  ✅ Пост #{pid} отклонён.{RESET}"); pause()

def restore_post():
    header("♻️  Восстановить отклонённый пост")
    with get_db() as db:
        rows = db.execute(
            """
            SELECT p.ID, p.URL, a.Name, r.RejectReason
            FROM posts_info p
            JOIN authors    a ON p.Author = a.ID
            LEFT JOIN results r ON r.Post = p.ID
            WHERE p.Status = 'rejected'
            ORDER BY p.ID DESC LIMIT 20
            """
        ).fetchall()
    if not rows: print(f"  {G}Нет отклонённых постов.{RESET}"); pause(); return
    opts = [f"#{r['ID']} {r['Name']} — {(r['RejectReason'] or '')[:30]}" for r in rows]
    idx  = menu("Выберите пост для восстановления", opts)
    if idx == -1: return
    pid = rows[idx]["ID"]
    with get_db() as db:
        db.execute("UPDATE posts_info SET Status='pending' WHERE ID=?", (pid,))
        db.execute("UPDATE results SET HumanWords=NULL, HumanErrors=NULL, Reviewer=NULL, RejectReason=NULL WHERE Post=?", (pid,))
        db.commit()
    if _QM:
        with get_db() as db:
            in_q = db.execute("SELECT 1 FROM queue WHERE Post=?", (pid,)).fetchone()
        if not in_q: assign_post(pid)
    print(f"\n{G}  ✅ Пост #{pid} восстановлен и возвращён в очередь.{RESET}"); pause()

def clear_queue():
    header("🗑  Очистить все очереди")
    total = get_total_queue_count() if _QM else 0
    if not confirm(f"Удалить ВСЕ {total} постов из очереди? (посты останутся со статусом pending)"): return
    with get_db() as db:
        db.execute("UPDATE posts_info SET Status='pending' WHERE Status IN ('checking', 'pending')")
        db.execute("DELETE FROM queue")
        db.commit()
    print(f"\n{G}  ✅ Все очереди очищены.{RESET}"); pause()

def reassign_queue():
    header("🔄 Переназначить очередь")
    with get_db() as db:
        count = db.execute(
            "SELECT COUNT(*) FROM posts_info WHERE Status IN ('pending', 'checking')"
        ).fetchone()[0]
    print(f"  Непроверенных постов: {BOLD}{count}{RESET}\n")
    if not confirm(f"Переназначить все {count} постов?"): return
    with get_db() as db:
        db.execute("UPDATE posts_info SET Status='pending' WHERE Status='checking'")
        db.execute("DELETE FROM queue")
        db.commit()
        posts = db.execute("SELECT ID FROM posts_info WHERE Status='pending' ORDER BY ID").fetchall()
    if not _QM: print(f"  {R}queue_manager недоступен.{RESET}"); pause(); return
    for p in posts: assign_post(p["ID"])
    print(f"\n{G}  ✅ Переназначено: {len(posts)} постов.{RESET}"); pause()


# ══════════════════════════════════════════════════════════════════════════════
# ДНИ
# ══════════════════════════════════════════════════════════════════════════════

def manage_days():
    while True:
        choice = menu("📅 Управление днями", [
            "Список дней", "Создать новый день",
            "Завершить и начать новый", "Удалить день",
        ])
        if choice == -1: return
        [list_days, create_new_day, finish_day, remove_day][choice]()

def list_days():
    header("📅 Дни конкурса")
    days = get_all_days()
    if not days:
        print(f"  {DIM}Дней нет.{RESET}")
    else:
        current = days[0]["day"] if days else None
        for d in days:
            marker = f" {C}← текущий{RESET}" if d["day"] == current else ""
            print(f"  День {d['day']}: {d['data']}{marker}")
    pause()

def create_new_day():
    header("➕ Создать новый день")
    from zoneinfo import ZoneInfo
    today = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")
    label = input(f"  Название/дата [{today}]: {W}").strip(); print(RESET, end="")
    if not label: label = today
    if not confirm(f"Создать день «{label}»?"): return
    create_day(label)
    print(f"\n{G}  ✅ День «{label}» создан.{RESET}"); pause()

def finish_day():
    header("🏁 Завершить текущий день")
    days = get_all_days()
    if not days: print(f"  {R}Нет активного дня.{RESET}"); pause(); return
    current = days[0]
    print(f"  Текущий день: {BOLD}{current['data']}{RESET}")
    if not confirm("Завершить и начать новый?"): return
    label = create_day()
    print(f"\n{G}  ✅ День {current['data']} завершён. Начат новый: {label}{RESET}"); pause()

def remove_day():
    header("🗑  Удалить день")
    days = get_all_days()
    if not days: print(f"  {R}Дней нет.{RESET}"); pause(); return
    opts = [f"День {d['day']}: {d['data']}" for d in days]
    idx  = menu("Выберите день для удаления", opts)
    if idx == -1: return
    day_id   = days[idx]["day"]
    day_data = days[idx]["data"]
    with get_db() as db:
        post_count = db.execute("SELECT COUNT(*) FROM posts_info WHERE Day=?", (day_id,)).fetchone()[0]
    transfer_to = None
    if post_count > 0:
        print(f"\n  {Y}⚠  К этому дню привязано {post_count} постов.{RESET}")
        other = [d for d in days if d["day"] != day_id]
        if not other: print(f"  {R}Нет других дней для переноса.{RESET}"); pause(); return
        t_idx = menu("Куда перенести посты?", [f"День {d['day']}: {d['data']}" for d in other])
        if t_idx == -1: return
        transfer_to = other[t_idx]["day"]
    if not confirm(f"Удалить день {day_id} ({day_data})?"): return
    delete_day(day_id, transfer_to)
    print(f"\n{G}  ✅ День {day_id} удалён.{RESET}"); pause()


# ══════════════════════════════════════════════════════════════════════════════
# ССЫЛКИ
# ══════════════════════════════════════════════════════════════════════════════

def manage_links():
    while True:
        choice = menu("🔗 Управление ссылками", [
            "Статистика ссылок", "Просмотр ссылок",
            "Удалить ссылку", "Добавить ссылку вручную",
            "Добавить в блэклист", "Просмотр блэклиста", "Удалить из блэклиста",
        ])
        if choice == -1: return
        [links_stats, view_links, delete_link, add_link,
         add_to_blacklist, view_blacklist, remove_from_blacklist][choice]()

def links_stats():
    header("🔗 Статистика ссылок")
    with get_db() as db:
        total    = db.execute("SELECT COUNT(*) FROM links").fetchone()[0]
        parsed   = db.execute("SELECT COUNT(*) FROM links WHERE Parsed=1").fetchone()[0]
        bl_count = db.execute("SELECT COUNT(*) FROM blacklist").fetchone()[0]
    print(f"  Всего ссылок:       {total}")
    print(f"  Распарсено:         {G}{parsed}{RESET}")
    print(f"  Ожидают парсинга:   {Y}{total - parsed}{RESET}")
    print(f"  В блэклисте:        {R}{bl_count}{RESET}")
    pause()

def view_links():
    header("🔗 Ссылки (последние 30)")
    with get_db() as db:
        rows = db.execute("SELECT URL, COALESCE(Parsed, 0) as Parsed FROM links ORDER BY rowid DESC LIMIT 30").fetchall()
    if not rows: print(f"  {DIM}Ссылок нет.{RESET}")
    else:
        for r in rows:
            status = f"{G}✅{RESET}" if r["Parsed"] else f"{Y}⏳{RESET}"
            print(f"  {status} {DIM}{(r['URL'] or '')[-80:]}{RESET}")
    pause()

def delete_link():
    header("🗑  Удалить ссылку")
    url = input(f"  URL или часть: {W}").strip(); print(RESET, end="")
    if not url: return
    with get_db() as db:
        rows = db.execute("SELECT URL FROM links WHERE URL LIKE ? LIMIT 10", (f"%{url}%",)).fetchall()
    if not rows: print(f"  {R}Не найдено.{RESET}"); pause(); return
    idx = menu("Выберите ссылку", [(r["URL"] or "")[-80:] for r in rows])
    if idx == -1: return
    if not confirm("Удалить ссылку?"): return
    with get_db() as db:
        db.execute("DELETE FROM links WHERE URL=?", (rows[idx]["URL"],)); db.commit()
    print(f"\n{G}  ✅ Ссылка удалена.{RESET}"); pause()

def add_link():
    header("➕ Добавить ссылку")
    url = input(f"  URL: {W}").strip(); print(RESET, end="")
    if not url: return
    with get_db() as db:
        db.execute("INSERT OR IGNORE INTO links (URL, Parsed) VALUES (?, 0)", (url,)); db.commit()
    print(f"\n{G}  ✅ Ссылка добавлена.{RESET}"); pause()

def add_to_blacklist():
    header("🚫 Добавить в блэклист")
    url = input(f"  URL: {W}").strip(); print(RESET, end="")
    if not url: return
    with get_db() as db:
        db.execute("INSERT OR IGNORE INTO blacklist (URL) VALUES (?)", (url,))
        db.execute("DELETE FROM links WHERE URL=?", (url,)); db.commit()
    print(f"\n{G}  ✅ Добавлено в блэклист.{RESET}"); pause()

def view_blacklist():
    header("🚫 Блэклист")
    with get_db() as db:
        rows = db.execute("SELECT URL FROM blacklist ORDER BY rowid DESC LIMIT 30").fetchall()
    if not rows: print(f"  {DIM}Блэклист пуст.{RESET}")
    else:
        for r in rows: print(f"  {DIM}{r['URL']}{RESET}")
    pause()

def remove_from_blacklist():
    header("♻️  Удалить из блэклиста")
    with get_db() as db:
        rows = db.execute("SELECT rowid, URL FROM blacklist ORDER BY rowid DESC LIMIT 20").fetchall()
    if not rows: print(f"  {DIM}Блэклист пуст.{RESET}"); pause(); return
    idx = menu("Выберите URL", [(r["URL"] or "")[-70:] for r in rows])
    if idx == -1: return
    with get_db() as db:
        db.execute("DELETE FROM blacklist WHERE rowid=?", (rows[idx]["rowid"],)); db.commit()
    print(f"\n{G}  ✅ Удалено из блэклиста.{RESET}"); pause()


# ══════════════════════════════════════════════════════════════════════════════
# ЛОГИ
# ══════════════════════════════════════════════════════════════════════════════

def view_logs():
    header("📄 Просмотр логов")
    logs_dir = ROOT / "logs"
    if not logs_dir.exists(): print(f"  {R}Папка логов не найдена.{RESET}"); pause(); return
    log_files = sorted(logs_dir.glob("*.log"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not log_files: print(f"  {R}Лог файлы не найдены.{RESET}"); pause(); return
    log_file = log_files[0]
    idx = menu(f"Файл: {log_file.name}\nСколько строк?", ["50 строк", "100 строк", "200 строк"])
    if idx == -1: return
    n = [50, 100, 200][idx]
    try:
        with open(log_file, "r", encoding="utf-8") as f:
            lines = f.readlines()
        last = lines[-n:] if len(lines) >= n else lines
    except Exception as e:
        print(f"  {R}Ошибка чтения: {e}{RESET}"); pause(); return
    header(f"📄 Последние {n} строк — {log_file.name}")
    for line in last:
        line = line.rstrip()
        if "ERROR" in line: print(f"  {R}{line}{RESET}")
        elif "WARNING" in line: print(f"  {Y}{line}{RESET}")
        elif "INFO" in line: print(f"  {DIM}{line}{RESET}")
        else: print(f"  {line}")
    pause()


# ══════════════════════════════════════════════════════════════════════════════
# ИНИЦИАЛИЗАЦИЯ БД
# ══════════════════════════════════════════════════════════════════════════════

SCHEMA = """
PRAGMA journal_mode = WAL;
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS reviewers (
    TGID     TEXT    PRIMARY KEY,
    URL      TEXT    NOT NULL UNIQUE,
    Name     TEXT    NOT NULL,
    IsAdmin  INTEGER NOT NULL DEFAULT 0,
    Verified INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS authors (
    ID   INTEGER PRIMARY KEY AUTOINCREMENT,
    Name TEXT    NOT NULL,
    URL  TEXT    NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS days (
    Day  INTEGER PRIMARY KEY AUTOINCREMENT,
    Data TEXT    NOT NULL
);

CREATE TABLE IF NOT EXISTS links (
    URL    TEXT    PRIMARY KEY,
    Parsed INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS blacklist (
    URL TEXT PRIMARY KEY
);

CREATE TABLE IF NOT EXISTS posts_info (
    ID     INTEGER PRIMARY KEY AUTOINCREMENT,
    Author INTEGER NOT NULL REFERENCES authors(ID),
    URL    TEXT    NOT NULL UNIQUE,
    Day    INTEGER REFERENCES days(Day),
    Status TEXT    NOT NULL DEFAULT 'pending'
        CHECK(Status IN ('pending','checking','done','rejected','reviewer_post'))
);

CREATE TABLE IF NOT EXISTS queue (
    Post       INTEGER NOT NULL PRIMARY KEY REFERENCES posts_info(ID),
    Reviewer   TEXT    REFERENCES reviewers(TGID),
    AssignedAt TEXT    NOT NULL DEFAULT (datetime('now','utc')),
    TakenAt    TEXT
);

CREATE TABLE IF NOT EXISTS results (
    ID           INTEGER PRIMARY KEY AUTOINCREMENT,
    Post         INTEGER NOT NULL UNIQUE REFERENCES posts_info(ID),
    BotWords     INTEGER,
    HumanWords   INTEGER,
    HumanErrors  INTEGER,
    RejectReason TEXT,
    Reviewer     TEXT REFERENCES reviewers(TGID)
);

CREATE INDEX IF NOT EXISTS idx_links_parsed     ON links(Parsed);
CREATE INDEX IF NOT EXISTS idx_posts_status     ON posts_info(Status);
CREATE INDEX IF NOT EXISTS idx_queue_reviewer   ON queue(Reviewer);
CREATE INDEX IF NOT EXISTS idx_queue_assignedat ON queue(AssignedAt);
CREATE INDEX IF NOT EXISTS idx_queue_takenat    ON queue(TakenAt);
CREATE INDEX IF NOT EXISTS idx_results_reviewer ON results(Reviewer);
"""

DB_PATH = ROOT / "data" / "main.db"

def init_db():
    header("🗄  Инициализация базы данных")
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    is_new = not DB_PATH.exists()
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.executescript(SCHEMA); conn.commit()
    finally:
        conn.close()
    status = f"{G}создана{RESET}" if is_new else f"{Y}уже существует, проверена{RESET}"
    print(f"  ✅ База данных {status}: {DB_PATH}\n")
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    tables = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
    ).fetchall()
    print(f"  {'Таблица':<22} {'Записей':>8}"); hr()
    for t in tables:
        count = conn.execute(f"SELECT COUNT(*) FROM {t['name']}").fetchone()[0]
        print(f"  {t['name']:<22} {count:>8}")
    conn.close(); pause()


# ══════════════════════════════════════════════════════════════════════════════
# ЭКСПОРТ
# ══════════════════════════════════════════════════════════════════════════════

def export_results():
    header("📊 Экспорт результатов в Excel")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print(f"  {R}❌ Не установлен openpyxl: pip install openpyxl{RESET}"); pause(); return

    def _font(size=10, bold=False, color="1A1A1A"):
        return Font(name="Arial", size=size, bold=bold, color=color)
    def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def _border():
        s = Side(style="thin", color="BDC3C7")
        return Border(left=s, right=s, top=s, bottom=s)
    def _align(h="center"): return Alignment(horizontal=h, vertical="center", wrap_text=True)
    def _hrow(ws, row, cols, bg="2C3E50"):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.font=_font(10,True,"FFFFFF"); cell.fill=_fill(bg)
            cell.alignment=_align(); cell.border=_border()
        ws.row_dimensions[row].height=22
    def _drow(ws, row, cols, even=False):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.font=_font(10); cell.fill=_fill("EBF5FB" if even else "FFFFFF")
            cell.alignment=_align("left" if c==1 else "center"); cell.border=_border()
        ws.row_dimensions[row].height=20

    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    days = conn.execute("SELECT Day, Data FROM days ORDER BY Day").fetchall()

    def posts_by_day(day_id):
        return conn.execute(
            """SELECT a.Name AS author, COUNT(p.ID) AS post_count,
                      COALESCE(SUM(r.HumanWords),0) AS words,
                      COALESCE(SUM(r.HumanErrors),0) AS errors
               FROM posts_info p JOIN authors a ON p.Author=a.ID JOIN results r ON r.Post=p.ID
               WHERE p.Day=? AND p.Status='done' AND r.HumanWords IS NOT NULL
               GROUP BY a.ID ORDER BY errors*1.0/NULLIF(words,0) ASC""", (day_id,)
        ).fetchall()

    reviewer_stats = conn.execute(
        """SELECT rv.Name, COUNT(r.ID) AS checked,
                  COALESCE(SUM(r.HumanWords),0) AS words,
                  COALESCE(SUM(r.HumanErrors),0) AS errors
           FROM reviewers rv
           LEFT JOIN results r ON r.Reviewer=rv.TGID AND r.HumanWords IS NOT NULL
           WHERE rv.Verified=1 GROUP BY rv.TGID ORDER BY checked DESC"""
    ).fetchall()

    top_authors = conn.execute(
        """SELECT a.Name AS author, COUNT(p.ID) AS post_count,
                  COALESCE(SUM(r.HumanWords),0) AS words,
                  COALESCE(SUM(r.HumanErrors),0) AS errors
           FROM posts_info p JOIN authors a ON p.Author=a.ID JOIN results r ON r.Post=p.ID
           WHERE p.Status='done' AND r.HumanWords IS NOT NULL
           GROUP BY a.ID HAVING words>0 ORDER BY errors*1.0/words ASC"""
    ).fetchall()

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Общий отчёт"); ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1"); ws["A1"].value="ИТОГИ КОНКУРСА inkstory.net"
    ws["A1"].font=_font(14,True,"FFFFFF"); ws["A1"].fill=_fill("2C3E50")
    ws["A1"].alignment=_align("center"); ws.row_dimensions[1].height=32

    ws.merge_cells("A2:E2")
    ws["A2"].value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font=_font(10,color="FFFFFF"); ws["A2"].fill=_fill("5D6D7E")
    ws["A2"].alignment=_align("center")

    r = 4
    for c,h in enumerate(["Жюри","Проверено","Слов","Ошибок","Ош/1000"],1):
        ws.cell(row=r,column=c).value=h
    _hrow(ws, r, 5, "3498DB")
    for i, rv in enumerate(reviewer_stats, 1):
        r+=1
        ep = f"=IFERROR(ROUND(D{r}/C{r}*1000,1),0)"
        for c,v in enumerate([rv["Name"],rv["checked"],rv["words"],rv["errors"],ep],1):
            ws.cell(row=r,column=c).value=v
        _drow(ws, r, 5, i%2==0)

    r+=2; ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"].value="ТОП УЧАСТНИКОВ — меньше всего ошибок на 1000 слов"
    ws[f"A{r}"].font=_font(11,True,"FFFFFF"); ws[f"A{r}"].fill=_fill("3498DB")
    ws[f"A{r}"].alignment=_align("center"); ws.row_dimensions[r].height=24

    r+=1
    for c,h in enumerate(["#","Участник","Постов","Слов","Ошибок","Ош/1000"],1):
        ws.cell(row=r,column=c).value=h
    _hrow(ws, r, 6, "3498DB")
    medals={1:"🥇",2:"🥈",3:"🥉"}
    for i,a in enumerate(top_authors,1):
        r+=1
        ep=f"=IFERROR(ROUND(E{r}/D{r}*1000,1),0)"
        for c,v in enumerate([medals.get(i,i),a["author"],a["post_count"],a["words"],a["errors"],ep],1):
            ws.cell(row=r,column=c).value=v
        _drow(ws, r, 6, i%2==0)

    for col,w in zip("ABCDE",[6,24,10,12,10]): ws.column_dimensions[col].width=w

    for day in days:
        ws_d = wb.create_sheet(day["Data"]); ws_d.sheet_view.showGridLines=False
        ws_d.merge_cells("A1:E1"); ws_d["A1"].value=day["Data"]
        ws_d["A1"].font=_font(13,True,"FFFFFF"); ws_d["A1"].fill=_fill("2C3E50")
        ws_d["A1"].alignment=_align("center"); ws_d.row_dimensions[1].height=28
        for c,h in enumerate(["Участник","Постов","Слов","Ошибок","Ош/1000"],1):
            ws_d.cell(row=3,column=c).value=h
        _hrow(ws_d, 3, 5, "3498DB")
        posts = posts_by_day(day["Day"])
        if not posts:
            ws_d.merge_cells("A4:E4"); ws_d["A4"].value="Нет данных"
            ws_d["A4"].alignment=_align("center")
        else:
            for i,p in enumerate(posts,1):
                rr=3+i; ep=f"=IFERROR(ROUND(D{rr}/C{rr}*1000,1),0)"
                for c,v in enumerate([p["author"],p["post_count"],p["words"],p["errors"],ep],1):
                    ws_d.cell(row=rr,column=c).value=v
                _drow(ws_d, rr, 5, i%2==0)
        for col,w in zip("ABCDE",[24,10,12,12,10]): ws_d.column_dimensions[col].width=w

    conn.close()
    results_dir = ROOT / "results"; results_dir.mkdir(exist_ok=True)
    out = results_dir / f"results_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    wb.save(out)
    print(f"  {G}✅ Экспорт завершён:{RESET} {out}"); pause()


# ══════════════════════════════════════════════════════════════════════════════
# ПОЛНАЯ ОЧИСТКА
# ══════════════════════════════════════════════════════════════════════════════

def full_reset():
    header("🗑  Полная очистка данных")
    TABLES = [("queue","Очередь"),("results","Результаты"),("posts_info","Посты"),("links","Ссылки"),("authors","Авторы")]
    with get_db() as db:
        print(f"  {'Таблица':<15} {'Записей':>8}"); hr()
        counts = {}
        for table, label in TABLES:
            count = db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            counts[table] = count
            print(f"  {label:<15} {R}{count:>8}{RESET}")
    print(f"\n  {R}{BOLD}⚠  Это удалит ВСЕ данные из этих таблиц!{RESET}")
    print(f"  {DIM}Таблицы reviewers, days и blacklist не затрагиваются.{RESET}\n")
    if not confirm("Ты уверен? Это действие необратимо"): print(f"\n  {DIM}Отменено.{RESET}"); pause(); return
    if not confirm("Подтверди ещё раз — удалить все данные"):   print(f"\n  {DIM}Отменено.{RESET}"); pause(); return
    with get_db() as db:
        for table, label in TABLES:
            db.execute(f"DELETE FROM {table}")
            print(f"  {G}✅ {label} очищена{RESET}  ({counts[table]} записей)")
        db.execute("DELETE FROM sqlite_sequence WHERE name IN ('results','posts_info','authors')")
        db.commit()
    print(f"\n{G}  ✅ Очистка завершена.{RESET}"); pause()


# ══════════════════════════════════════════════════════════════════════════════
# ГЛАВНОЕ МЕНЮ
# ══════════════════════════════════════════════════════════════════════════════

def main():
    actions = [
        show_stats, manage_reviewers, manage_posts,
        manage_days, manage_links, view_logs,
        init_db, export_results, full_reset,
    ]
    try:
        while True:
            choice = menu("Главное меню", [
                "📊 Статистика",
                "👥 Управление жюри",
                "📝 Управление постами",
                "📅 Управление днями",
                "🔗 Управление ссылками",
                "📄 Просмотр логов",
                "🗄  Инициализация БД",
                "📤 Экспорт результатов",
                "🗑  Полная очистка данных",
            ], show_stats=True)
            if choice == -1: break
            actions[choice]()
    except KeyboardInterrupt: pass
    finally: clr(); print(f"{DIM}Выход.{RESET}\n")

if __name__ == "__main__":
    main()