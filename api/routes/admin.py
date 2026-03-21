"""
api/routes/admin.py — эндпоинты только для администраторов

GET  /api/admin/reviewers                   — список жюри
POST /api/admin/reviewers/{tgid}/verify     — верифицировать
POST /api/admin/reviewers/{tgid}/unverify   — снять верификацию
POST /api/admin/reviewers/{tgid}/make-admin — назначить админом
POST /api/admin/reviewers/{tgid}/remove-admin — снять права
DELETE /api/admin/reviewers/{tgid}          — удалить жюри

GET  /api/admin/queue        — очереди по жюри
GET  /api/admin/posts        — посты с фильтром по статусу
GET  /api/admin/days         — список дней
POST /api/admin/days         — создать день
DELETE /api/admin/days/{day} — удалить день

GET  /api/admin/stats        — расширенная статистика
GET  /api/admin/logs         — последние строки лога
GET  /api/admin/export       — экспорт Excel (возвращает файл)
"""

import io
import pathlib
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from api.auth import get_admin_user
from utils.db.jury import (
    get_all_reviewers, get_reviewer_stats,
    set_verified, set_admin, delete_reviewer,
)
from utils.db.posts import get_posts_stats, errors_per_1000
from utils.db.days import get_all_days, create_day, delete_day
from utils.database import get_db
from utils.constants import PostStatus
from parser.queue_manager import get_all_reviewer_queue_sizes, get_total_queue_count
from utils.logger import setup_logger

log      = setup_logger()
router   = APIRouter(prefix="/admin", tags=["admin"])
LOG_FILE = pathlib.Path(__file__).parent.parent.parent / "logs" / "app.log"


# ── Схемы ─────────────────────────────────────────────────────────────────────

class CreateDay(BaseModel):
    label: str | None = Field(None, max_length=50)


# ── Жюри ──────────────────────────────────────────────────────────────────────

@router.get("/reviewers")
def list_reviewers(user: dict = Depends(get_admin_user)):
    return get_all_reviewers()


@router.post("/reviewers/{tgid}/verify")
def verify_reviewer(tgid: str, user: dict = Depends(get_admin_user)):
    set_verified(tgid, 1)
    log.info(f"[api/admin] Верифицирован {tgid} (by {user['tg_id']})")
    # Обновляем кнопку Mini App для нового жюри
    try:
        from utils.config import BOT_TOKEN
        import requests as _req
        import pathlib, os
        from dotenv import load_dotenv
        load_dotenv(pathlib.Path(__file__).parent.parent.parent / ".env")
        tunnel_url = os.getenv("_TUNNEL_URL", "")
        if tunnel_url:
            _req.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/setChatMenuButton",
                json={
                    "chat_id": tgid,
                    "menu_button": {
                        "type": "web_app",
                        "text": "Открыть панель",
                        "web_app": {"url": tunnel_url},
                    }
                },
                timeout=5,
            )
    except Exception:
        pass
    return {"ok": True}


@router.post("/reviewers/{tgid}/unverify")
def unverify_reviewer(tgid: str, user: dict = Depends(get_admin_user)):
    set_verified(tgid, 0)
    log.info(f"[api/admin] Снята верификация {tgid} (by {user['tg_id']})")
    return {"ok": True}


@router.post("/reviewers/{tgid}/make-admin")
def make_admin(tgid: str, user: dict = Depends(get_admin_user)):
    set_admin(tgid, 1)
    log.info(f"[api/admin] Назначен админ {tgid} (by {user['tg_id']})")
    return {"ok": True}


@router.post("/reviewers/{tgid}/remove-admin")
def remove_admin(tgid: str, user: dict = Depends(get_admin_user)):
    set_admin(tgid, 0)
    log.info(f"[api/admin] Сняты права админа {tgid} (by {user['tg_id']})")
    return {"ok": True}


@router.delete("/reviewers/{tgid}")
def remove_reviewer(tgid: str, user: dict = Depends(get_admin_user)):
    delete_reviewer(tgid)
    log.info(f"[api/admin] Удалён жюри {tgid} (by {user['tg_id']})")
    return {"ok": True}


# ── Очередь ───────────────────────────────────────────────────────────────────

@router.get("/queue")
def get_queue(user: dict = Depends(get_admin_user)):
    rows = get_all_reviewer_queue_sizes()
    return {
        "total":     get_total_queue_count(),
        "reviewers": rows,
    }


# ── Посты ─────────────────────────────────────────────────────────────────────

@router.get("/posts")
def get_posts(
    status: str | None = Query(None, description="pending|checking|done|rejected"),
    limit: int = Query(50, ge=1, le=200),
    user: dict = Depends(get_admin_user),
):
    allowed = {None} | set(PostStatus.ALL)
    if status not in allowed:
        raise HTTPException(status_code=400, detail=f"Invalid status: {status}")

    where = "" if not status else f"WHERE p.Status = '{status}'"

    with get_db() as db:
        rows = db.execute(
            f"""
            SELECT p.ID, p.URL, p.Status, a.Name AS author,
                   r.BotWords, r.HumanWords, r.HumanErrors, r.RejectReason,
                   rv.Name AS reviewer_name
            FROM posts_info p
            JOIN authors a ON p.Author = a.ID
            LEFT JOIN results   r  ON r.Post     = p.ID
            LEFT JOIN reviewers rv ON rv.TGID    = r.Reviewer
            {where}
            ORDER BY p.ID DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()

    return [
        {
            "id":            r["ID"],
            "url":           r["URL"],
            "status":        r["Status"],
            "author":        r["author"],
            "bot_words":     r["BotWords"],
            "human_words":   r["HumanWords"],
            "human_errors":  r["HumanErrors"],
            "errors_per_1000": errors_per_1000(r["HumanErrors"] or 0, r["HumanWords"] or 0),
            "reject_reason": r["RejectReason"],
            "reviewer":      r["reviewer_name"],
        }
        for r in rows
    ]


# ── Дни ───────────────────────────────────────────────────────────────────────

@router.get("/days")
def list_days(user: dict = Depends(get_admin_user)):
    return get_all_days()


@router.post("/days")
def add_day(body: CreateDay, user: dict = Depends(get_admin_user)):
    label = create_day(body.label)
    log.info(f"[api/admin] Создан день '{label}' (by {user['tg_id']})")
    return {"ok": True, "label": label}


@router.delete("/days/{day_id}")
def remove_day(
    day_id: int,
    transfer_to: int | None = Query(None),
    user: dict = Depends(get_admin_user),
):
    delete_day(day_id, transfer_to)
    log.info(f"[api/admin] Удалён день {day_id} (by {user['tg_id']})")
    return {"ok": True}


# ── Статистика ────────────────────────────────────────────────────────────────

@router.get("/stats")
def get_admin_stats(user: dict = Depends(get_admin_user)):
    stats     = get_posts_stats()
    reviewers = get_reviewer_stats()
    return {
        **stats,
        "in_queue":  get_total_queue_count(),
        "reviewers": [
            {
                **r,
                "errors_per_1000": errors_per_1000(
                    r.get("total_errors", 0), r.get("total_words", 0)
                ),
            }
            for r in reviewers
        ],
    }


# ── Логи ──────────────────────────────────────────────────────────────────────

@router.get("/logs")
def get_logs(
    n: int = Query(100, ge=10, le=500),
    user: dict = Depends(get_admin_user),
):
    try:
        with open(LOG_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
        last = lines[-n:] if len(lines) >= n else lines
        return {"lines": "".join(last)}
    except FileNotFoundError:
        return {"lines": ""}


# ── Экспорт Excel ─────────────────────────────────────────────────────────────

@router.get("/export")
async def export_excel(user: dict = Depends(get_admin_user)):
    """Генерирует Excel-отчёт и отправляет файл в Telegram личкой."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    CLR_HEADER = "2C3E50"
    CLR_ACCENT = "3498DB"
    CLR_EVEN   = "EBF5FB"
    CLR_WHITE  = "FFFFFF"
    CLR_LIGHT  = "FFFFFF"
    CLR_DARK   = "1A1A1A"

    def _font(size=10, bold=False, color=CLR_DARK):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        s = Side(style="thin", color="BDC3C7")
        return Border(left=s, right=s, top=s, bottom=s)

    def _align(h="center"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def _header_row(ws, row, cols, bg=CLR_HEADER):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font      = _font(10, bold=True, color=CLR_LIGHT)
            cell.fill      = _fill(bg)
            cell.alignment = _align("center")
            cell.border    = _border()
        ws.row_dimensions[row].height = 22

    def _data_row(ws, row, cols, even=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font      = _font(10)
            cell.fill      = _fill(CLR_EVEN if even else CLR_WHITE)
            cell.alignment = _align("left" if c == 1 else "center")
            cell.border    = _border()
        ws.row_dimensions[row].height = 20

    with get_db() as db:
        days = db.execute("SELECT Day, Data FROM days ORDER BY Day").fetchall()

        def get_posts_by_day(day_id):
            return db.execute(
                """
                SELECT a.Name AS author, COUNT(p.ID) AS post_count,
                       COALESCE(SUM(r.HumanWords), 0) AS words,
                       COALESCE(SUM(r.HumanErrors), 0) AS errors
                FROM posts_info p
                JOIN authors a ON p.Author = a.ID
                JOIN results r ON r.Post = p.ID
                WHERE p.Day = ? AND p.Status = 'done' AND r.HumanWords IS NOT NULL
                GROUP BY a.ID
                ORDER BY errors * 1.0 / NULLIF(words, 0) ASC
                """, (day_id,)
            ).fetchall()

        reviewer_stats = db.execute(
            """
            SELECT rv.Name, COUNT(r.ID) AS checked,
                   COALESCE(SUM(r.HumanWords), 0) AS words,
                   COALESCE(SUM(r.HumanErrors), 0) AS errors
            FROM reviewers rv
            LEFT JOIN results r ON r.Reviewer = rv.TGID AND r.HumanWords IS NOT NULL
            WHERE rv.Verified = 1
            GROUP BY rv.TGID ORDER BY checked DESC
            """
        ).fetchall()

        top_authors = db.execute(
            """
            SELECT a.Name AS author, COUNT(p.ID) AS post_count,
                   COALESCE(SUM(r.HumanWords), 0) AS words,
                   COALESCE(SUM(r.HumanErrors), 0) AS errors
            FROM posts_info p
            JOIN authors a ON p.Author = a.ID
            JOIN results r ON r.Post = p.ID
            WHERE p.Status = 'done' AND r.HumanWords IS NOT NULL
            GROUP BY a.ID HAVING words > 0
            ORDER BY errors * 1.0 / words ASC
            """
        ).fetchall()

        wb = Workbook()
        wb.remove(wb.active)

        # Лист: Общий отчёт
        ws = wb.create_sheet("Общий отчёт")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:E1")
        ws["A1"].value     = "ИТОГИ КОНКУРСА inkstory.net"
        ws["A1"].font      = _font(14, bold=True, color=CLR_LIGHT)
        ws["A1"].fill      = _fill(CLR_HEADER)
        ws["A1"].alignment = _align("center")

        ws.merge_cells("A2:E2")
        ws["A2"].value     = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A2"].font      = _font(10, color=CLR_LIGHT)
        ws["A2"].fill      = _fill("5D6D7E")
        ws["A2"].alignment = _align("center")

        # Жюри
        r = 4
        for c, h in enumerate(["Жюри", "Проверено", "Слов", "Ошибок", "Ош/1000"], 1):
            ws.cell(row=r, column=c).value = h
        _header_row(ws, r, 5, CLR_ACCENT)

        for i, rv in enumerate(reviewer_stats, 1):
            r += 1
            ep = f"=IFERROR(ROUND(D{r}/C{r}*1000,1),0)"
            for c, v in enumerate([rv["Name"], rv["checked"], rv["words"], rv["errors"], ep], 1):
                ws.cell(row=r, column=c).value = v
            _data_row(ws, r, 5, even=(i % 2 == 0))

        # Топ авторов
        r += 2
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value     = "ТОП УЧАСТНИКОВ"
        ws[f"A{r}"].font      = _font(11, bold=True, color=CLR_LIGHT)
        ws[f"A{r}"].fill      = _fill(CLR_ACCENT)
        ws[f"A{r}"].alignment = _align("center")

        r += 1
        for c, h in enumerate(["Участник", "Постов", "Слов", "Ошибок", "Ош/1000"], 1):
            ws.cell(row=r, column=c).value = h
        _header_row(ws, r, 5, CLR_ACCENT)

        medals = {1: "🥇", 2: "🥈", 3: "🥉"}
        for i, a in enumerate(top_authors, 1):
            r += 1
            ep = f"=IFERROR(ROUND(D{r}/C{r}*1000,1),0)"
            for c, v in enumerate([f"{medals.get(i,i)} {a['author']}", a["post_count"], a["words"], a["errors"], ep], 1):
                ws.cell(row=r, column=c).value = v
            _data_row(ws, r, 5, even=(i % 2 == 0))

        for col, w in zip("ABCDE", [24, 10, 12, 12, 10]):
            ws.column_dimensions[col].width = w

        # Листы по дням
        for day in days:
            ws_d = wb.create_sheet(day["Data"])
            ws_d.sheet_view.showGridLines = False
            ws_d.merge_cells("A1:E1")
            ws_d["A1"].value     = day["Data"]
            ws_d["A1"].font      = _font(13, bold=True, color=CLR_LIGHT)
            ws_d["A1"].fill      = _fill(CLR_HEADER)
            ws_d["A1"].alignment = _align("center")

            for c, h in enumerate(["Участник", "Постов", "Слов", "Ошибок", "Ош/1000"], 1):
                ws_d.cell(row=3, column=c).value = h
            _header_row(ws_d, 3, 5, CLR_ACCENT)

            posts = get_posts_by_day(day["Day"])
            if not posts:
                ws_d.merge_cells("A4:E4")
                ws_d["A4"].value     = "Нет данных"
                ws_d["A4"].alignment = _align("center")
            else:
                for i, p in enumerate(posts, 1):
                    row = 3 + i
                    ep  = f"=IFERROR(ROUND(D{row}/C{row}*1000,1),0)"
                    for c, v in enumerate([p["author"], p["post_count"], p["words"], p["errors"], ep], 1):
                        ws_d.cell(row=row, column=c).value = v
                    _data_row(ws_d, row, 5, even=(i % 2 == 0))

            for col, w in zip("ABCDE", [24, 10, 12, 12, 10]):
                ws_d.column_dimensions[col].width = w

    # Сохраняем и отправляем в Telegram
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"results_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Также сохраняем в папку results/
    import pathlib
    results_dir = pathlib.Path(__file__).parent.parent.parent / "results"
    results_dir.mkdir(exist_ok=True)
    out_path = results_dir / filename
    with open(out_path, "wb") as f:
        f.write(buf.getvalue())

    # Отправляем файл в Telegram
    try:
        from utils.config import BOT_TOKEN
        import httpx
        buf.seek(0)
        async with httpx.AsyncClient() as client:
            await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                data={"chat_id": user["tg_id"], "caption": f"📊 Экспорт результатов\n{filename}"},
                files={"document": (filename, buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30,
            )
    except Exception as e:
        log.error(f"[export] Ошибка отправки в Telegram: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка отправки: {e}")

    return {"ok": True, "filename": filename}
